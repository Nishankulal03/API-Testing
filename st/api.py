import os
import json
from flask import Flask, request, flash, redirect, url_for, render_template
from werkzeug.utils import secure_filename
import openpyxl
import requests

app = Flask(__name__)
app.secret_key = '61c140ba91bbfb981732dd873c055a5840acc0e665dfbf6b'

# Define a folder to store uploaded files
UPLOAD_FOLDER = './uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Limit allowed file extensions to Excel files
ALLOWED_EXTENSIONS = {'xlsx'}

# Function to check if file extension is allowed
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Redirect root route to upload page
@app.route('/')
def index():
    return redirect(url_for('upload_file'))

# Route for uploading file and displaying upload form
@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # Check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        # If user does not select file, browser also submit an empty part without filename
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)
            flash(f'File successfully uploaded: {filename}')
            return redirect(url_for('run_tests', filename=filename))
        else:
            flash('File type not allowed')
            return redirect(request.url)
    return render_template('upload.html')

@app.route('/run_tests/<filename>')
def run_tests(filename):
    input_file = os.path.join(UPLOAD_FOLDER, filename)
    output_file = os.path.join(UPLOAD_FOLDER, 'output_' + filename)
    result_message = run_api_tests(input_file, output_file)
    if result_message.startswith('Error'):
        flash(result_message)
        return redirect(url_for('upload_file'))
    return f'Tests completed. Output saved as {output_file}.'

def run_api_tests(input_file, output_file):
    try:
        # Load input Excel file
        wb = openpyxl.load_workbook(input_file)
        sheet = wb.active

        # Create output Excel file
        output_wb = openpyxl.Workbook()
        output_sheet = output_wb.active

        # Set headers for output file
        headers = ["TestCaseID", "Method", "URL", "Headers", "Payload",
                   "ExpectedStatusCode", "ExpectedResponse", "Result", "ActualStatusCode", "ActualResponse"]
        output_sheet.append(headers)

        # Iterate through test cases
        for row in sheet.iter_rows(min_row=2, values_only=True):
            TestCaseID, Method, URL, Headers, Payload, ExpectedStatusCode, ExpectedResponse = row[:7]

            # Convert Headers from string to dict if not empty
            headers_dict = {}
            if Headers:
                headers_list = Headers.split('\n')
                headers_dict = dict(line.split(': ', 1) for line in headers_list)

            # Convert Payload from JSON string to dict if not empty
            payload_str = ''
            if Payload:
                try:
                    payload_dict = json.loads(Payload)  # Convert string to dict safely
                    payload_str = json.dumps(payload_dict, indent=4)  # Convert dict back to pretty string
                except json.JSONDecodeError as e:
                    payload_str = f"Error parsing Payload: {e}"

            # Make the API request
            try:
                if Method.upper() == 'GET':
                    response = requests.get(URL, headers=headers_dict)
                elif Method.upper() == 'POST':
                    headers_dict['Content-Type'] = 'application/json'
                    response = requests.post(URL, headers=headers_dict, json=json.loads(Payload))
                elif Method.upper() == 'PUT':
                    headers_dict['Content-Type'] = 'application/json'
                    response = requests.put(URL, headers=headers_dict, json=json.loads(Payload))
                elif Method.upper() == 'DELETE':
                    response = requests.delete(URL, headers=headers_dict)
                else:
                    raise ValueError(f"Unsupported HTTP Method: {Method}")

                ActualResponse = response.text
                ActualStatusCode = response.status_code

                # Parse JSON response if possible
                try:
                    json_response = response.json()
                except ValueError:
                    json_response = None

                # Compare with expected values
                if str(ActualStatusCode) == str(ExpectedStatusCode):
                    if ExpectedResponse and ExpectedResponse in ActualResponse:
                        Result = "Pass"
                    elif json_response and ExpectedResponse in json_response.values():
                        Result = "Pass"
                    elif not ExpectedResponse:  # Handle case where ExpectedResponse is None or empty
                        Result = "Pass"
                    else:
                        Result = "Fail"
                else:
                    Result = "Fail"

            except Exception as e:
                ActualResponse = str(e)
                ActualStatusCode = "Exception"
                Result = "Fail"

            # Append results to output Excel file
            output_sheet.append([TestCaseID, Method, URL, Headers, payload_str,
                                 ExpectedStatusCode, ExpectedResponse, Result, ActualStatusCode, ActualResponse])

        # Save the output Excel file
        output_wb.save(output_file)
        return 'Tests completed. Output saved as {output_file}.'

    except FileNotFoundError:
        return f"Error: File '{input_file}' not found."
    except Exception as e:
        return f"Error: {str(e)}"

if __name__ == '__main__':
    app.run(debug=True)
